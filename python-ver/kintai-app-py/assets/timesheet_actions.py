"""出退勤ロジック・Excel書込・CSV出力"""
import csv
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Callable, Dict, List, Optional, Any

from assets.app_logger import get_logger
_log = get_logger("kintai.actions")

try:
    import openpyxl
    from openpyxl.utils.datetime import to_excel
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from assets.timesheet_constants import (
    START_TIME_MAP, END_TIME_MAP, REALTIME_SHIFTS,
    VACATION_FIXED, VACATION_INPUT, HALF_DAY_PAID,
    VACATION_CONFIG, VACATION_INPUT_CONFIG,
    WORK_STYLE_REMOTE, WORK_STYLE_OFFICE
)
from assets.timesheet_helpers import (
    round_time, round_time_night_shift, time_to_excel_serial, find_timesheet,
    is_late, format_date_jp, get_row_for_date, get_now, get_today
)


class TimesheetNotFoundError(Exception):
    """タイムシートファイルが見つからない"""
    def __init__(self, folder: str, name: str, year: int, month: int):
        self.folder = folder
        self.name = name
        self.year = year
        self.month = month
        super().__init__(f"タイムシートが見つかりません: {name} {year}年{month:02d}月")


class TimesheetLockedError(Exception):
    """Excelファイルが別プロセスに開かれている"""
    def __init__(self, path):
        self.path = Path(path)
        super().__init__(f"ファイルが別のプロセスに開かれています: {self.path.name}")


class TimesheetWriteError(Exception):
    """その他のExcel書込エラー"""
    pass


class UnknownShiftTypeError(Exception):
    """出勤形態に対応する処理が定義されていない"""
    def __init__(self, shift: str):
        self.shift = shift
        super().__init__(f"「{shift}」の処理はまだ定義されていません")


def clock_in(
    config,
    shift: str,
    work_style: str,
    target_date: date,
    is_assumed: bool,
    no_post: bool,
    late_reason_cb: Callable,
    half_day_cb: Callable,
    remark_cb: Callable,
    status_cb: Callable,
) -> bool:
    """
    出勤処理。
    status_cb(msg, color) でUIへフィードバック。
    Returns True on success, False on failure.
    """
    _log.info("clock_in 開始: date=%s shift=%s work_style=%s is_assumed=%s", target_date, shift, work_style, is_assumed)
    try:
        if shift in VACATION_FIXED:
            # 固定休暇: VACATION_CONFIG から各列の値を取得
            config_entry = VACATION_CONFIG[shift]
            start_serial = None
            end_serial = None
            if config_entry["start_time"]:
                sh, sm = map(int, config_entry["start_time"].split(":"))
                start_serial = time_to_excel_serial(sh, sm)
            if config_entry["end_time"]:
                eh, em = map(int, config_entry["end_time"].split(":"))
                end_serial = time_to_excel_serial(eh, em)
            row_data = {
                "date": target_date,
                "shift_label": config_entry["shift_label"],
                "start_time": start_serial,
                "end_time": end_serial,
                "overtime_type": None,
                "remark": config_entry["remark"],
            }

        elif shift in VACATION_INPUT:
            # 備考入力が必要な休暇: VACATION_INPUT_CONFIG からダイアログプロンプトを取得
            input_config = VACATION_INPUT_CONFIG[shift]
            result = remark_cb(input_config["dialog_prompt"])
            if result is None:
                status_cb("キャンセルされました", "gray")
                return False, ""
            remark = result
            row_data = {
                "date": target_date,
                "shift_label": input_config["shift_label"],
                "start_time": None,
                "end_time": None,
                "overtime_type": None,
                "remark": remark,
            }

        elif shift == HALF_DAY_PAID:
            # 0.5日有給
            result = half_day_cb()
            if result is None:
                status_cb("キャンセルされました", "gray")
                return False, ""
            start_q = result["start"]
            end_q = result["end"]
            remark = result.get("remark", "")
            start_serial = time_to_excel_serial(start_q.hour(), start_q.minute())
            end_serial = time_to_excel_serial(end_q.hour(), end_q.minute())
            row_data = {
                "date": target_date,
                "shift_label": "0.5日有給",
                "start_time": start_serial,
                "end_time": end_serial,
                "overtime_type": None,
                "remark": remark if remark else None,
            }

        elif shift in REALTIME_SHIFTS:
            # リアルタイムシフト
            now = get_now()

            if is_assumed:
                # 想定記入: E列=出勤形態, F列=固定始業, G列=固定終業
                start_str = START_TIME_MAP[shift]
                end_str = END_TIME_MAP[shift]
                sh, sm = map(int, start_str.split(":"))
                eh, em = map(int, end_str.split(":"))
                row_data = {
                    "date": target_date,
                    "shift_label": shift,
                    "start_time": time_to_excel_serial(sh, sm),
                    "end_time": time_to_excel_serial(eh, em),
                    "overtime_type": None,
                    "remark": None,
                }
            elif not is_assumed and is_late(shift, now):
                # 遅刻: E列="遅刻", F列=丸め後実時刻, L列=遅刻理由
                reason = late_reason_cb()
                if reason is None:
                    status_cb("キャンセルされました", "gray")
                    return False, ""
                if shift == "深夜":
                    rounded = round_time_night_shift(now)
                    start_serial = time_to_excel_serial(rounded["hours"], rounded["minutes"])
                else:
                    rounded_dt = round_time(now)
                    start_serial = time_to_excel_serial(rounded_dt.hour, rounded_dt.minute)
                row_data = {
                    "date": target_date,
                    "shift_label": "遅刻",
                    "start_time": start_serial,
                    "end_time": None,
                    "overtime_type": None,
                    "remark": reason,
                }
            else:
                # 通常出勤: E列=出勤形態, F列=固定始業
                start_str = START_TIME_MAP[shift]
                sh, sm = map(int, start_str.split(":"))
                row_data = {
                    "date": target_date,
                    "shift_label": shift,
                    "start_time": time_to_excel_serial(sh, sm),
                    "end_time": None,
                    "overtime_type": None,
                    "remark": None,
                }

        else:
            raise UnknownShiftTypeError(shift)

        # CSV出力判定
        if not is_assumed and shift in REALTIME_SHIFTS and target_date == get_today():
            output_csv(config, shift, work_style, target_date)

        # Teams投稿（Excel書込より先に実行）
        # リアルタイムシフト AND 非想定 AND 本日のみ投稿
        late_reason = row_data.get("remark") if row_data.get("shift_label") == "遅刻" else ""
        time_str = ""
        if row_data.get("start_time") is not None:
            serial = row_data["start_time"]
            minutes_total = round(serial * 24 * 60)
            h, m = divmod(minutes_total, 60)
            time_str = f"{h:02d}:{m:02d}"
        teams_error = ""
        if not no_post and not is_assumed and shift in REALTIME_SHIFTS and target_date == get_today():
            try:
                from assets.teams_webhook import send_teams_post
                send_teams_post(config, "clock_in", {
                    "shift": shift,
                    "work_style": work_style,
                    "comment": row_data.get("remark") or "",
                })
            except Exception as e:
                teams_error = f"Teams投稿エラー: {e}"

        # Excel書込（エラーは呼び出し元に伝播させてダイアログ表示）
        xlsx_path = _find_xlsx_or_raise(config, target_date)
        if xlsx_path:
            write_to_excel(xlsx_path, row_data)

        _log.info("clock_in 完了: date=%s shift=%s teams_error=%s", target_date, shift, teams_error or "なし")
        return True, teams_error

    except (TimesheetNotFoundError, TimesheetLockedError, TimesheetWriteError, UnknownShiftTypeError) as e:
        _log.warning("clock_in エラー: %s", e)
        raise
    except Exception as e:
        _log.error("clock_in 予期しないエラー: %s", e, exc_info=True)
        status_cb(f"出勤処理エラー: {e}", "red")
        return False, ""


def clock_out(
    config,
    shift: str,
    work_style: str,
    target_date: date,
    no_post: bool,
    clock_out_info: dict,
    status_cb: Callable,
    is_cross_day: bool = False,
) -> bool:
    """
    退勤処理。
    clock_out_info keys: next_workday(date), next_shift(str), mention(str), comment(str)
    is_cross_day: 退勤時刻が日付を跨いでいる場合 True
    """
    _log.info("clock_out 開始: shift=%s is_cross_day=%s", shift, is_cross_day)
    try:
        now = get_now()

        # ターゲット日付決定
        is_night = (shift == "深夜")
        if is_cross_day:
            if is_night:
                actual_target = now - timedelta(days=2)
            else:
                actual_target = now - timedelta(days=1)
        else:
            if is_night:
                actual_target = now - timedelta(days=1)
            else:
                actual_target = now
        target_date = actual_target.date() if isinstance(actual_target, datetime) else actual_target

        # 時刻丸め
        # ・通常シフト 通常退勤: 実時刻そのまま（24h表記）
        # ・それ以外（深夜通常/通常日跨ぎ/深夜日跨ぎ）: 実時刻 + 24h
        #   ※深夜日跨ぎ（翌々日退勤）も +24h のみ。前前日行に書くことで区別する。
        rounded_dt = round_time(now)
        if is_cross_day or is_night:
            end_serial = time_to_excel_serial(rounded_dt.hour + 24, rounded_dt.minute)
        else:
            end_serial = time_to_excel_serial(rounded_dt.hour, rounded_dt.minute)

        # 残業判定: F列(始業時刻)を読み取り G列との差分で9h超を判定
        overtime_type = None
        xlsx_path = _find_xlsx_or_raise(config, target_date)
        if xlsx_path and OPENPYXL_AVAILABLE:
            try:
                wb_check = openpyxl.load_workbook(str(xlsx_path))
                ws_check = wb_check.active
                row_num_check = get_row_for_date(ws_check, target_date.day)
                if row_num_check is None:
                    raise TimesheetWriteError(
                        f"「{xlsx_path.name}」に{target_date.month}月{target_date.day}日の行を認識できませんでした。"
                    )
                start_value = ws_check.cell(row=row_num_check, column=6).value
                if start_value is None or start_value == "":
                    # 始業時間が記載されていない場合はエラー
                    raise TimesheetWriteError("始業時間が記載されていません。先に出勤を記録してください。")
                # openpyxl は時刻書式セルを datetime.time / datetime.datetime で返す場合がある
                if isinstance(start_value, (int, float)):
                    start_serial = float(start_value)
                elif hasattr(start_value, 'hour') and hasattr(start_value, 'minute'):
                    start_serial = (start_value.hour * 60 + start_value.minute) / (24 * 60)
                else:
                    start_serial = float(start_value)
                nine_hours_serial = 9.0 / 24.0
                work_duration = end_serial - start_serial
                if work_duration > nine_hours_serial:
                    overtime_type = "客先指示"
                wb_check.close()
            except (TimesheetWriteError, TimesheetLockedError):
                raise
            except Exception:
                pass

        # 備考設定
        remark = clock_out_info.get("comment", "") or None

        row_data = {
            "date": target_date,
            "shift_label": None,   # 退勤時はE列を上書きしない
            "start_time": None,    # 退勤時はF列を上書きしない
            "end_time": end_serial,
            "overtime_type": overtime_type,
            "remark": remark,
        }

        # Teams投稿（Excel書込より先に実行）
        # リアルタイムシフトのみ投稿
        # Teams表示用は実時刻（24h+なし）
        time_str = f"{rounded_dt.hour:02d}:{rounded_dt.minute:02d}"

        # 実労働時間（Teams投稿用）
        start_str = START_TIME_MAP.get(shift, "10:00")
        sh, sm = map(int, start_str.split(":"))
        start_serial_est = time_to_excel_serial(sh, sm)
        hours_worked = (end_serial - start_serial_est) * 24

        teams_error = ""
        if not no_post and shift in REALTIME_SHIFTS:
            try:
                from assets.teams_webhook import send_teams_post
                send_teams_post(config, "clock_out", {
                    "next_workday": clock_out_info.get("next_workday"),
                    "next_shift": clock_out_info.get("next_shift", ""),
                    "next_work_mode": clock_out_info.get("next_work_mode", ""),
                    "mention": clock_out_info.get("mention", ""),
                    "comment": clock_out_info.get("comment", ""),
                })
            except Exception as e:
                teams_error = f"Teams投稿エラー: {e}"

        # Excel書込（エラーは呼び出し元に伝播させてダイアログ表示）
        if xlsx_path:
            write_to_excel(xlsx_path, row_data)

        _log.info("clock_out 完了: date=%s shift=%s overtime=%s teams_error=%s",
                  target_date, shift, overtime_type or "なし", teams_error or "なし")
        return True, teams_error

    except (TimesheetNotFoundError, TimesheetLockedError, TimesheetWriteError) as e:
        _log.warning("clock_out エラー: %s", e)
        raise
    except Exception as e:
        _log.error("clock_out 予期しないエラー: %s", e, exc_info=True)
        status_cb(f"退勤処理エラー: {e}", "red")
        return False, ""


def batch_write(
    config,
    dates: List[date],
    shift: str,
    work_style: str,
    half_day_cb: Callable,
    remark_cb: Callable,
    status_cb: Callable,
) -> tuple:
    """複数日付のループ記入（一括記入）。(success_count, fail_count) を返す。"""
    _log.info("batch_write 開始: shift=%s dates=%d件", shift, len(dates))
    success_count = 0
    fail_count = 0

    # 未定義の出勤形態チェック（ループ前に検証）
    _known = set(VACATION_FIXED) | set(VACATION_INPUT) | {HALF_DAY_PAID} | set(REALTIME_SHIFTS)
    if shift not in _known:
        raise UnknownShiftTypeError(shift)

    # 0.5日有給・VACATION_INPUT の場合は最初に一度だけ入力を求める
    shared_half_day = None
    shared_remark = None

    if shift == HALF_DAY_PAID:
        shared_half_day = half_day_cb()
        if shared_half_day is None:
            status_cb("キャンセルされました", "gray")
            return 0, 0
    elif shift in VACATION_INPUT:
        input_config = VACATION_INPUT_CONFIG[shift]
        shared_remark = remark_cb(input_config["dialog_prompt"])
        if shared_remark is None:
            status_cb("キャンセルされました", "gray")
            return 0, 0

    for d in dates:
        try:
            if shift in VACATION_FIXED:
                config_entry = VACATION_CONFIG[shift]
                start_serial = None
                end_serial = None
                if config_entry["start_time"]:
                    sh, sm = map(int, config_entry["start_time"].split(":"))
                    start_serial = time_to_excel_serial(sh, sm)
                if config_entry["end_time"]:
                    eh, em = map(int, config_entry["end_time"].split(":"))
                    end_serial = time_to_excel_serial(eh, em)
                row_data = {
                    "date": d,
                    "shift_label": config_entry["shift_label"],
                    "start_time": start_serial,
                    "end_time": end_serial,
                    "overtime_type": None,
                    "remark": config_entry["remark"],
                }
            elif shift in VACATION_INPUT:
                input_config = VACATION_INPUT_CONFIG[shift]
                row_data = {
                    "date": d,
                    "shift_label": input_config["shift_label"],
                    "start_time": None,
                    "end_time": None,
                    "overtime_type": None,
                    "remark": shared_remark,
                }
            elif shift == HALF_DAY_PAID:
                start_q = shared_half_day["start"]
                end_q = shared_half_day["end"]
                remark = shared_half_day.get("remark", "")
                start_serial = time_to_excel_serial(start_q.hour(), start_q.minute())
                end_serial = time_to_excel_serial(end_q.hour(), end_q.minute())
                row_data = {
                    "date": d,
                    "shift_label": "0.5日有給",
                    "start_time": start_serial,
                    "end_time": end_serial,
                    "overtime_type": None,
                    "remark": remark if remark else None,
                }
            elif shift in REALTIME_SHIFTS:
                start_str = START_TIME_MAP[shift]
                end_str = END_TIME_MAP[shift]
                sh, sm = map(int, start_str.split(":"))
                eh, em = map(int, end_str.split(":"))
                row_data = {
                    "date": d,
                    "shift_label": shift,
                    "start_time": time_to_excel_serial(sh, sm),
                    "end_time": time_to_excel_serial(eh, em),
                    "overtime_type": None,
                    "remark": None,
                }
            else:
                raise UnknownShiftTypeError(shift)

            # Excel書込（エラーはステータスラベルに表示して続行）
            try:
                xlsx_path = _find_xlsx_or_raise(config, d)
                if xlsx_path:
                    write_to_excel(xlsx_path, row_data)
            except TimesheetNotFoundError as e:
                _log.warning("batch_write タイムシート未検出: date=%s %s", d, e)
                status_cb(
                    f"{d.strftime('%Y/%m/%d')} タイムシートが見つかりませんでした。", "orange"
                )
                fail_count += 1
                continue
            except TimesheetLockedError as e:
                _log.warning("batch_write ファイルロック: date=%s %s", d, e)
                status_cb(
                    f"{d.strftime('%Y/%m/%d')} Excelが開かれています: {e.path.name}", "orange"
                )
                fail_count += 1
                continue
            except TimesheetWriteError as e:
                _log.warning("batch_write 書込エラー: date=%s %s", d, e)
                status_cb(f"{d.strftime('%Y/%m/%d')} Excel書込エラー: {e}", "orange")
                fail_count += 1
                continue

            _log.debug("batch_write 書込成功: date=%s", d)
            success_count += 1

        except Exception as e:
            _log.error("batch_write 予期しないエラー: date=%s %s", d, e, exc_info=True)
            fail_count += 1
            status_cb(f"{d.strftime('%Y/%m/%d')} エラー: {e}", "orange")

    _log.info("batch_write 完了: 成功=%d 失敗=%d", success_count, fail_count)
    return success_count, fail_count


def write_to_excel(xlsx_path: Path, row_data: dict) -> bool:
    """
    openpyxl で Excel に書込む。
    C列を走査して target_day に一致する行を特定し、各列に値を書込む。

    row_data keys:
      - date: date (対象日)
      - shift_label: str or None (E列: 就労テキスト "日勤","遅刻" 等)
      - start_time: float or None (F列: 始業Excelシリアル値)
      - end_time: float or None (G列: 終業Excelシリアル値)
      - overtime_type: str or None (K列: 残業種別 "客先指示" 等)
      - remark: str or None (L列: 備考)
    """
    if not OPENPYXL_AVAILABLE:
        return False

    try:
        target_date = row_data["date"]

        wb = openpyxl.load_workbook(str(xlsx_path))
        ws = wb.active

        row_num = get_row_for_date(ws, target_date.day)
        if row_num is None:
            raise TimesheetWriteError(
                f"「{xlsx_path.name}」に{target_date.month}月{target_date.day}日の行を認識できませんでした。"
            )

        # E列(5): 就労（出勤形態テキスト）
        if row_data.get("shift_label") is not None:
            ws.cell(row=row_num, column=5).value = row_data["shift_label"]

        # F列(6): 始業時刻（Excelシリアル値）
        if row_data.get("start_time") is not None:
            ws.cell(row=row_num, column=6).value = row_data["start_time"]

        # G列(7): 終業時刻（Excelシリアル値）
        if row_data.get("end_time") is not None:
            ws.cell(row=row_num, column=7).value = row_data["end_time"]

        # K列(11): 残業種別
        if row_data.get("overtime_type") is not None:
            ws.cell(row=row_num, column=11).value = row_data["overtime_type"]

        # L列(12): 備考
        if row_data.get("remark") is not None:
            ws.cell(row=row_num, column=12).value = row_data["remark"]

        wb.save(str(xlsx_path))
        _log.debug("write_to_excel 完了: file=%s date=%s row=%d", xlsx_path.name, target_date, row_num)
        return True

    except PermissionError:
        raise TimesheetLockedError(xlsx_path)
    except Exception as e:
        raise TimesheetWriteError(f"Excel書込エラー: {e}")


def output_csv(
    config,
    shift: str,
    work_style: str,
    target_date: date,
) -> None:
    """
    CSV出力 ({name}.csv) 上書き
    フォーマット: name,shift
    name  : shift_display_name（シフト表上の名前表記）
    shift : シフト種別。リモートの場合は末尾に "(ﾃ" を付与（例: 日勤(ﾃ）
    """
    try:
        output_dir = Path(config.output_folder if config else "attendance_data")
        output_dir.mkdir(parents=True, exist_ok=True)

        name = getattr(config, "shift_display_name", "") or (config.display_name if config else "")
        csv_path = output_dir / f"{name}.csv"
        shift_text = f"{shift}(ﾃ" if work_style == WORK_STYLE_REMOTE else shift

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["name", "shift"])
            writer.writerow([name, shift_text])
    except Exception:
        pass


def _find_xlsx_or_raise(config, target_date: date) -> Path:
    """
    タイムシートxlsxを検索して返す。
    フォルダ未設定・名前未設定・ファイル未検出のいずれでも
    TimesheetNotFoundError を raise する。
    """
    folder = (config.timesheet_folder if config else "") or ""
    name = ""
    if config:
        name = getattr(config, "timesheet_display_name", "") or config.display_name or ""

    if not folder:
        raise TimesheetNotFoundError(
            "(タイムシートフォルダ未設定)", name or "(未設定)",
            target_date.year, target_date.month
        )
    if not name:
        raise TimesheetNotFoundError(
            folder, "(タイムシートの名前表記が未設定)",
            target_date.year, target_date.month
        )

    xlsx = find_timesheet(folder, name, target_date.year, target_date.month)
    if not xlsx:
        raise TimesheetNotFoundError(folder, name, target_date.year, target_date.month)
    return xlsx
